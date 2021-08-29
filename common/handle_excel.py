# -*- coding: utf-8 -*-
# @Time    : 2021/8/29 9:05
# @Author  : kanghe
# @Email   : 244783726@qq.com
# @File    : handle_excel.py

from openpyxl import load_workbook

from common import handle_path as project
from common.handle_log import log


def load_all_data_from_xls(filename, sheet_name: str, parent=project.casedatas_dir):
    """ 获取 excel 中指定 sheet 中的所有数据
    注意:
        1 标题行也就是首行不能为空
        2 不会读取整行都为空的数据

    :param filename: excel 文件路径
    :param sheet_name: 需要读取的 sheet 索引，从 0 开始，默认为 0（即默认为第一个 sheet）
    :param parent: 文件所在的文件夹
    :return: 返回一个 list of dict 格式的数据，存储所有的数据
    """
    try:
        wb = load_workbook(parent.joinpath(filename), read_only=True)
        sheet = wb[sheet_name.strip()]
    except FileNotFoundError as e:
        log.error(f'excel文件不存在---{e}')
        return f'excel文件不存在---{e}'
    except IndexError as e:
        log.error(f'sheet不存在---{e}')
        return f'sheet不存在---{e}'

    sheet_values = []

    title = [j.value for i in sheet.iter_rows(max_row=1) for j in i if j.value]
    for i in sheet.iter_rows(max_col=len(title), min_row=2):
        tmp = [j.value for j in i]
        if any(tmp):
            sheet_values.append(dict(zip(title, tmp)))
    return sheet_values


if __name__ == '__main__':

    print(load_all_data_from_xls("api_cases.xlsx", "注册"))
